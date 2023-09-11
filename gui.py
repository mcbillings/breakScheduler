##import files/modules
from employee import *
from breakScheduler import *
from tkinter import *
from tkinter import ttk, simpledialog
from tkinter.ttk import Progressbar
from Pmw import ScrolledFrame
from tkinter import filedialog
import csv
import xlsxwriter
from openpyxl import *
import re

##variable/list declairations
##Employee arrays
empObjs = []
breakObjs = []

date = ""

def insertionSort(arr):
    ##Using insertion sort since employees in a given day will be around 21-55
    ##will also assign each employee to their department since we're already iterating over
    ##every item
    for i in range(1, len(arr)):
        ##Sorting Stuff
        key = arr[i].getStartTime()
        keyObj = arr[i]
        
        j = i-1
        while j >= 0 and key < arr[j].getStartTime():
            arr[j + 1] = arr[j]
            j -= 1
        arr[j + 1] = keyObj

##Break Objs by department
mBreaks = breakSlots(0, "Manager", 0, 0)
cBreaks = breakSlots(53, "Cashier", 7, 0)
spBreaks = breakSlots(35, "Soft Production", 0, 6)
ssBreaks = breakSlots(53, "Soft Stocking", 11, 6)
hpBreaks = breakSlots(35, "Hard Production", 6, 6)
hsBreaks = breakSlots(53, "Hard Stocking" , 19, 6)
accBreaks = breakSlots(53, "ACC", 16, 0)
shBreaks = breakSlots(35, "Shoes", 22, 0)
njBreaks = breakSlots(35, "New Goods/Jewlery", 24, 0)
bkBreaks = breakSlots(35, "Books", 26, 0)
wBreaks = breakSlots(35, "Warehouse", 31, 6)
oBreaks = breakSlots(53, "Other", 28, 6)

##dictionary mapping lists to departments
departments = {"Cashier": cBreaks,
               "Soft Production": spBreaks,
               "Soft Stocking": ssBreaks,
               "Hard Production": hpBreaks,
               "Hard Stocking": hsBreaks,
               "ACC": accBreaks,
               "Shoes": shBreaks,
               "New Goods/Jewlery": njBreaks,
               "Books": bkBreaks,
               "Warehouse": wBreaks,
               "Other": oBreaks,
               "Fitting Room": oBreaks,
               "Ecom": oBreaks}

#################################
##----------ROOT SETUP---------##
#################################
root = Tk()
root.title("Break Scheduler V1.0.3")
root.state("iconic")

##getting number of emps
numEmps = simpledialog.askinteger("Enter number of Employees", "Enter number of Employees")
root.state("normal")

#Grid configuration
Grid.columnconfigure(root, 0, weight = 4)
Grid.columnconfigure(root, 1, weight = 3)

Grid.rowconfigure(root, 0, weight = 2)
Grid.rowconfigure(root, 1, weight = 15)
Grid.rowconfigure(root, 2, weight = 3)
Grid.rowconfigure(root, 3, weight = 1)


#################################
##--------LABEL SETUP----------##
#################################

Label(root, text = "Break Scheduler", font = ("Arial", 20)).grid(row = 0, column = 0, sticky = 'w')
Label(root, text = "Date: ", font = ("Arial", 20)).grid(row = 0, column = 1, sticky = 'w')

#################################
##--------EMPLOYEE SETUP-------##
#################################

#Employee arrays used to populate employee objects after being entered
empsNames = []
empsDepts = []
empsStart = []
empsEnd = []
empsSPM = []
empsEPM = []
empsBackup = []
empsSpecial = []

##function to add input fields
def addEmp():
    ##i used to position newest field in the correct row
    i = len(empsNames)
    
    ##updating numEmps for use when scheduling
    global numEmps
    numEmps += 1
    
    ##Adding name field
    empName = Entry(empsFrame)
    empsNames.append(empName)
    empName.grid(row = i+1, column = 0, sticky = 'ew')
    
    ##adding department dropdown
    empDept = StringVar()
    empDepartment = OptionMenu(empsFrame, empDept, *departmentsList)
    empsDepts.append(empDept)
    empDepartment.grid(row = i+1, column = 1, sticky = 'ew')
    
    
    ##adding start time field
    empStart = Entry(empsFrame, width = 10)
    empsStart.append(empStart)
    empStart.grid(row = i+1, column = 2, sticky = 'ew')
    
    ##Adding start PM checkbox
    s = IntVar(value = 0)
    empSPM = Checkbutton(empsFrame, variable = s)
    empsSPM.append(s)
    empSPM.grid(row = i + 1, column = 3)
    
    
    ##adding end time field
    empEnd = Entry(empsFrame, width = 10)
    empsEnd.append(empEnd)
    empEnd.grid(row = i+1, column = 4, sticky = 'ew')
    
    
    ##adding end PM checkbox
    e = IntVar(value = 1)
    empEPM = Checkbutton(empsFrame, variable = e)
    empsEPM.append(e)
    empEPM.grid(row = i + 1, column = 5)
    
    ##Adding backup checkbox
    b = IntVar(value = 0)
    empBackup = Checkbutton(empsFrame, variable = b)
    empsBackup.append(b)
    empBackup.grid(row = i + 1, column = 6)
    
    ##Adding special attribute dropdown
    empSpec = StringVar()
    empSpecial = OptionMenu(empsFrame, empSpec, *specialList)
    empsSpecial.append(empSpec)
    empSpecial.grid(row = i+1, column = 7, sticky = 'ew')
    
    ##Drawing to screen
    scrollableFrame.grid(row = 1, column = 0, sticky = 'nsew')

#Frame for all input fields
scrollableFrame = ScrolledFrame(root, horizflex = 'expand')
empsFrame = scrollableFrame.interior()


#Lists to be used with dropdown menus
departmentsList = ["Manager",
                   "Cashier",
                   "Soft Production",
                   "Soft Stocking",
                   "Hard Production",
                   "Hard Stocking",
                   "ACC",
                   "Shoes",
                   "New Goods/Jewlery",
                   "Books",
                   "Warehouse",
                   "Fitting Room",
                   "Other"]

specialList = ["None",
               "Pricer",
               "Clean",
               "Nytch",
               "Ecom"]

#Grid configuration
Grid.columnconfigure(empsFrame, index = 0, weight = 9)
Grid.columnconfigure(empsFrame, index = 1, weight = 12)
Grid.columnconfigure(empsFrame, index = 2, weight = 2)
Grid.columnconfigure(empsFrame, index = 4, weight = 2)
Grid.columnconfigure(empsFrame, index = 7, weight = 5)


for i in range(numEmps):
    #Populating input area with input fields
    addEmp()
    numEmps -= 1
    
#Labels for input fields
Label(empsFrame, text = "Name").grid(row = 0, column = 0, sticky = 'w')
Label(empsFrame, text = "Department").grid(row = 0, column = 1, sticky = 'w')
Label(empsFrame, text = "Start").grid(row = 0, column = 2, sticky = 'w')
Label(empsFrame, text = "PM?").grid(row = 0, column = 3, sticky = 'w')
Label(empsFrame, text = "End").grid(row = 0, column = 4, sticky = 'w')
Label(empsFrame, text = "PM?").grid(row = 0, column = 5, sticky = 'w')
Label(empsFrame, text = "Backup?").grid(row = 0, column = 6, sticky = 'w')
Label(empsFrame, text = "Special Attribute").grid(row = 0, column = 7, sticky = 'w')
  
#binding mousewheel to employee input frame
def _on_mousewheel(event):
    scrollableFrame.yview(mode = 'scroll', value = int(-1*event.delta/120))  

scrollableFrame.bind_all("<MouseWheel>", _on_mousewheel)

#Griding frame to screen
scrollableFrame.grid(row = 1, column = 0, sticky = 'nsew')


#################################
##--------SCHEDULE SETUP-------##
#################################
scheduleFrame = Frame(root)

Grid.columnconfigure(scheduleFrame, 0, weight = 6)
Grid.columnconfigure(scheduleFrame, 1, weight = 2)
Grid.columnconfigure(scheduleFrame, 2, weight = 2)
Grid.columnconfigure(scheduleFrame, 3, weight = 2)
Grid.columnconfigure(scheduleFrame, 4, weight = 2)
Grid.columnconfigure(scheduleFrame, 6, weight = 6)
Grid.columnconfigure(scheduleFrame, 7, weight = 2)
Grid.columnconfigure(scheduleFrame, 8, weight = 2)
Grid.columnconfigure(scheduleFrame, 9, weight = 2)
Grid.columnconfigure(scheduleFrame, 10, weight = 2)

def drawSchedule():
    for col in range(5):
        for row in range(33):
            Grid.rowconfigure(scheduleFrame, index = row, weight = 1)
            if col == 0 or col == 5:
                Label(scheduleFrame, borderwidth = 1, relief = 'solid').grid(row = row, column = col, sticky = 'nsew')
                Label(scheduleFrame, borderwidth = 1, relief = 'solid').grid(row = row, column = col + 6, sticky = 'nsew')
            else:
                Label(scheduleFrame, borderwidth = 1, relief = 'solid').grid(row = row, column = col, sticky = 'nsew')
                Label(scheduleFrame, borderwidth = 1, relief = 'solid').grid(row = row, column = col + 6, sticky = 'nsew')
            

    ##class labels
    Label(scheduleFrame, text = "Managers",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 0, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 1, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 2, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 3, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 4, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Cashiers",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 7, column = 0, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 7, column = 1, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 7, column = 2, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 7, column = 3, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 7, column = 4, sticky = 'nsew')
    
    Label(scheduleFrame, text = "ACC", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 16, column = 0, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 16, column = 1, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 16, column = 2, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 16, column = 3, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 16, column = 4, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Shoes", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 22, column = 0, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 22, column = 1, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 22, column = 2, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 22, column = 3, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 22, column = 4, sticky = 'nsew')
    
    Label(scheduleFrame, text = "New Goods/Jewlery", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 24, column = 0, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 24, column = 1, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 24, column = 2, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 24, column = 3, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 24, column = 4, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Books", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 26, column = 0, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 26, column = 1, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 26, column = 2, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 26, column = 3, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 26, column = 4, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Soft Production", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 6, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 7, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 8, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 9, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 0, column = 10, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Wares Production", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 6, column = 6, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 6, column = 7, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 6, column = 8, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 6, column = 9, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 6, column = 10, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Apparel Stockers", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 11, column = 6, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 11, column = 7, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 11, column = 8, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 11, column = 9, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 11, column = 10, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Wares Stockers", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 19, column = 6, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 19, column = 7, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 19, column = 8, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 19, column = 9, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 19, column = 10, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Other", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 28, column = 6, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 28, column = 7, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 28, column = 8, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 28, column = 9, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 28, column = 10, sticky = 'nsew')
    
    Label(scheduleFrame, text = "Warehouse", borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 31, column = 6, sticky = 'nsew')
    Label(scheduleFrame, text = "Shift",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 31, column = 7, sticky = 'nsew')
    Label(scheduleFrame, text = "1st Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 31, column = 8, sticky = 'nsew')
    Label(scheduleFrame, text = "Lunch",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 31, column = 9, sticky = 'nsew')
    Label(scheduleFrame, text = "2nd Break",  borderwidth = 1, relief = 'solid', font=('Helvetica', 11, 'bold')).grid(row = 31, column = 10, sticky = 'nsew')

    scheduleFrame.grid(row = 1, column = 1, sticky = 'nsew', padx = 10)
    
drawSchedule()

#################################
##--------OPTIONS SETUP--------##
#################################

optionsFrame = ttk.Frame(root)
    
Grid.columnconfigure(optionsFrame, 0, weight = 2)
Grid.columnconfigure(optionsFrame, 1, weight = 2)
Grid.columnconfigure(optionsFrame, 2, weight = 2)
Grid.columnconfigure(optionsFrame, 3, weight = 10)
Grid.columnconfigure(optionsFrame, 4, weight = 2)

def addDate():
    global date
    date = simpledialog.askstring("Enter Date", "Enter Date")
    Label(root, text = f'Date: {date}', font = ("Arial", 20)).grid(row = 0, column = 1, sticky = 'w')
    
def readEmps():
    filename = filedialog.askopenfilename(title = "Select a File",
                                          filetypes = (("XLSX Files",
                                                        "*.xlsx*"),
                                                       ("all files",
                                                        "*.*")))
    
    weekday = {"sunday": 3,
               "monday": 6,
               "tuesday": 9,
               "wednesday": 12,
               "thursday": 15,
               "friday": 18,
               "saturday": 21}

    depts = ["Manager",
             "Recovery",
             "New Goods/Jewlery",
             "Cashier",
             "Books",
             "Soft Production",
             "Soft Stocking",
             "Hard Production",
             "Hard Stocking",
             "Shoes",
             "ACC",
             "Fitting Room",
             "Warehouse",
             "Other"]

    altDepts = {"fr": "Fitting Room",
                "shoes": "Shoes",
                "books": "Books",
                "cash": "Cashier",
                "ss": "Soft Stocking",
                "acc": "ACC",
                "sp": "Soft Production",
                "hp": "Hard Production",
                "hs": "Hard Stocking",
                "wh": "Warehouse"}

    specAtts = {"ny": "Nytch",
                "pulls": "Pulls",
                "clean": "Clean",
                "nco": "NCO",
                "ecom": "Ecom"}

    day = simpledialog.askstring("Enter Day", "Enter Day of the Week")
    dayCol = weekday[day]
    deptID = 0
    deptColor = "FFEAD1DC"
    

    scheduleWB = load_workbook(filename, data_only=True)
    scheduleWS = scheduleWB.active
    
    global date
    date = str(scheduleWS.cell(row = 4, column = dayCol + 1).value)
    date = re.split('[- ]', date)
    print(date)
    date = date[1] + "-" + date[2]
    Label(root, text = f'Date: {date}', font = ("Arial", 20)).grid(row = 0, column = 1, sticky = 'w')

    for row in scheduleWS.iter_rows(max_row = 10):
        if row[0].value == "MANAGEMENT TEAM":
            startRow = row[0].row + 1
            break
    
    rowNum = 1

    for row in scheduleWS.iter_rows(min_row = startRow, max_row = 160):
        if row[0].fill.start_color.index == deptColor:
            ##change dept
            deptID += 1
        elif row[0].fill.start_color.index == "00000000" or "FFFFFFFF":
            if row[0].value is not None:
                ##employee found, set values
                thisEmpName = row[0].value
                thisEmpName = thisEmpName.split()
                thisEmpName = str(thisEmpName[0] + " " + thisEmpName[1][0])
                shiftStr = str(row[dayCol].internal_value)
                shift = re.split("(\d+)", shiftStr)
                thisEmpStart = 0
                thisEmpEnd = 0
                thisEmpDept = ""
                startPM = False
                thisEmpSpecAtt = ""
                if len(shift) == 1:
                    ##not an employee
                    thisEmpStart = -1
                elif len(shift) == 13:
                    #excel converted data to datetime, must account for this
                    thisEmpStart = int(shift[3]) #converting to int because it will have preceeding 0
                    thisEmpEnd = int(shift[5])
                    
                    empStartHP = toHP(thisEmpStart, True)
                    empEndHP = toHP(thisEmpEnd, True)
                    
                    if empStartHP - empEndHP < int(row[dayCol + 1].value):
                        startPM = True
                    
                    thisEmpDept = depts[deptID]
                elif len(shift[4]) == 0:
                    ##employee is in cur dept and has no special attribute
                    thisEmpStart = int(shift[1])
                    thisEmpEnd = int(shift[3])
                    
                    empStartHP = toHP(thisEmpStart, True)
                    empEndHP = toHP(thisEmpEnd, True)
                    
                    if empStartHP - empEndHP < int(row[dayCol + 1].value):
                        startPM = True
                        
                    thisEmpDept = depts[deptID]
                else:
                    ##employee is in different dept or has special attribute
                    thisEmpStart = int(shift[1])
                    thisEmpEnd = int(shift[3])
                    
                    empStartHP = toHP(thisEmpStart, True)
                    empEndHP = toHP(thisEmpEnd, True)
                    
                    if empStartHP - empEndHP < int(row[dayCol + 1].value):
                        startPM = True
                    
                    suffix = str(shift[4]).lower()
                    if suffix in specAtts:
                        ##don't change dept, just apply specAtts
                        thisEmpDept = depts[deptID]
                        thisEmpSpecAtt = specAtts[suffix]
                        if suffix == "ecom":
                            thisEmpDept = "Other"
                    elif suffix in altDepts:
                        thisEmpDept = altDepts[str(shift[4]).lower()]
                
                if thisEmpStart != -1:
                    print(str(thisEmpName) + ": " + str(thisEmpStart) + "-" + str(thisEmpEnd) + " " + thisEmpDept + " " + str(startPM))
                    addEmp()
                    ##Adding name field
                    empName = Entry(empsFrame)
                    empName.insert(rowNum, thisEmpName)
                    empsNames.append(empName)
                    empName.grid(row = rowNum, column = 0, sticky = 'ew')

                    ##adding department dropdown
                    empDept = StringVar()
                    empDepartment = OptionMenu(empsFrame, empDept, *departmentsList)
                    empDept.set(thisEmpDept)
                    empsDepts.append(empDept)
                    empDepartment.grid(row = rowNum, column = 1, sticky = 'ew')


                    ##adding start time field
                    empStart = Entry(empsFrame, width = 10)
                    empStart.insert(rowNum, thisEmpStart)
                    empsStart.append(empStart)
                    empStart.grid(row = rowNum, column = 2, sticky = 'ew')
                    
                    ##Adding start PM checkbox
                    if startPM:
                        thisVal = 0
                    else:
                        thisVal = 1
                    s = IntVar(value = startPM)
                    empSPM = Checkbutton(empsFrame, variable = s)
                    empsSPM.append(s)
                    empSPM.grid(row = rowNum, column = 3)


                    ##adding end time field
                    empEnd = Entry(empsFrame, width = 10)
                    empEnd.insert(rowNum, thisEmpEnd)
                    empsEnd.append(empEnd)
                    empEnd.grid(row = rowNum, column = 4, sticky = 'ew')
                    
                    ##adding end PM checkbox
                    e = IntVar(value = 1)
                    empEPM = Checkbutton(empsFrame, variable = e)
                    empsEPM.append(e)
                    empEPM.grid(row = rowNum, column = 5)
                    
                    ##Adding backup checkbox
                    b = IntVar(value = 0)
                    empBackup = Checkbutton(empsFrame, variable = b)
                    empsBackup.append(b)
                    empBackup.grid(row = rowNum, column = 6)
                    
                    ##Adding special attribute dropdown
                    empSpec = StringVar()
                    empSpecial = OptionMenu(empsFrame, empSpec, *specialList)
                    empSpec.set(thisEmpSpecAtt)
                    empsSpecial.append(empSpec)
                    empSpecial.grid(row = rowNum, column = 7, sticky = 'ew')
                    
                    rowNum = rowNum + 1
                    
                    global numEmps
                    numEmps = numEmps + 1
            
            
        scrollableFrame.grid(row = 1, column = 0, sticky = 'nsew')
        
def exportSchedule():
    directory = filedialog.askdirectory()
    
    # Create an new Excel file and add a worksheet.
    global date
    workbook = xlsxwriter.Workbook(directory+f'/break-schedule-{date}.xlsx')
    worksheet = workbook.add_worksheet()

    # Widen the first column to make the text clearer.
    worksheet.set_column('A:A', 20)
    worksheet.set_column('F:F', 2)
    worksheet.set_column('G:G', 20)
    
    labelFormat = workbook.add_format()
    labelFormat.set_bold()
    labelFormat.set_border()
    cleanFormat = workbook.add_format()
    cleanFormat.set_bg_color('#BCF3F9')
    cleanFormat.set_border()
    defaultFormat = workbook.add_format()
    defaultFormat.set_border()
    
    for child in scheduleFrame.children.values():
        row = child.grid_info()["row"]
        col = child.grid_info()["column"]
        worksheet.write(row + 1, col, child["text"])
        if child.cget("bg") == '#BCF3F9':
            worksheet.write(row + 1, col, child["text"], cleanFormat)
        elif child.cget("font") == 'Helvetica 11 bold':
            worksheet.write(row + 1, col, child["text"], labelFormat)
        else:
            worksheet.write(row + 1, col, child["text"], defaultFormat)
            
    worksheet.write(0, 0, f'Date: {date}')
            
        

    workbook.close()
    
            
            
Button(optionsFrame, text = "Add Employee", command = addEmp).grid(row = 1, column = 0)
Button(optionsFrame, text = "Schedule", command = lambda: initEmps(empsNames,
                                                                   empsDepts,
                                                                   empsStart,
                                                                   empsSPM,
                                                                   empsEnd,
                                                                   empsEPM,
                                                                   empsBackup,
                                                                   empsSpecial,
                                                                   breakObjs,
                                                                   empObjs)).grid(row = 1, column = 1)
Button(optionsFrame, text = "Import File", command = readEmps).grid(row = 1, column = 2)
Button(optionsFrame, text = "Add Date", command = addDate).grid(row = 1, column = 4)
Button(optionsFrame, text = "Export Schedule", command = exportSchedule).grid(row = 1, column = 3)

optionsFrame.grid(row = 2, column = 0, columnspan = 2, sticky='nsew', pady = 5)

def initEmps(names, depts, start, sPM, end, ePM, backup, special, breakObjs, empObjs):
    drawSchedule()
    for i in range(numEmps):
        print(names[i].get())
        if (len(names[i].get()) == 0 or len(start[i].get()) == 0 or len(end[i].get()) == 0 or len(depts[i].get()) == 0):
            pass
        else:
            ##converting Time
            startConverted = toHP(int(start[i].get()), int(sPM[i].get()))
            print("Start: " + str(startConverted))
            endConverted = toHP(int(end[i].get()), int(ePM[i].get()))
            print("End: " + str(endConverted))
            
            empObj = Employees(str(names[i].get()),
                               startConverted,
                               endConverted,
                               str(depts[i].get()),
                               special[i].get(),
                               backup[i].get())
            print(empObj.name + " has " + str(empObj.restBreaks) + " rest breaks")
            empObjs.append(empObj)

    insertionSort(empObjs)
    
    cIndex = 0
    
    for i in range(len(empObjs)):
        ##Adding to Department's Object array
        if empObjs[i].department == 'Manager':
            empObjs[i].index = len(mBreaks.emps)
            mBreaks.emps.append(empObjs[i])
        else:
            empDept = empObjs[i].getDepartment()
            deptClass = departments.get(empDept)
            
            if empDept == 'Cashier' and cIndex > 0:
                empObjs[i].index = len(deptClass.emps) - cIndex ##correct index for positioning
            else:
                empObjs[i].index = len(deptClass.emps) ##correct index for positioning
            
            
            if empObjs[i].getBackup() == 1:
                print("Adding to cBreaks")
                cBreaks.emps.append(empObjs[i])
                ##add a placeholder in dept so others don't get misplaced
                placeholder = Employees.placeholderEmp()
                deptClass.emps.append(placeholder)
                cIndex += 1
                
            else:
                deptClass.emps.append(empObjs[i])

    breakObjs=[mBreaks, cBreaks, spBreaks, ssBreaks, hpBreaks, hsBreaks, accBreaks,
               shBreaks, njBreaks, bkBreaks, wBreaks, oBreaks]
    
    for dept in range(len(breakObjs)):
        for emp in range(len(breakObjs[dept].emps)):
            empObj = breakObjs[dept].emps[emp]
                
            if empObj is not None:
                if empObj.department == "Manager":
                    empInd = empObj.index
                    empDispRow = mBreaks.dispRow
                    empDispCol = mBreaks.dispCol ##doing it this way so backups are displayed on their own row
                    empColor = '#F0F0F0'
                    
                    ##Not printing breaks if they are 0
                    if empObj.getRestBreak1() == 0:
                        b1 = ""
                    else:
                        b1 = empObj.getRestBreak1()
                    if empObj.getMealBreak() == 0:
                        b2 = ""
                    else:
                        b2 = empObj.getMealBreak()
                    if empObj.getRestBreak2() == 0:
                        b3 = ""
                    else:
                        b3 = empObj.getRestBreak2()
                    
                    Label(scheduleFrame, text = empObj.getName(), borderwidth = 1, relief = 'solid', bg = empColor).grid(row = empInd + empDispRow + 1,
                                                                                                              column = empDispCol,
                                                                                                              sticky = 'nsew')
                    Label(scheduleFrame, text = str(toHM(empObj.getStartTime())) + "-"+ str(toHM(empObj.getEndTime())),
                          borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                  column = empDispCol + 1,
                                                                  sticky = 'nsew')
                    Label(scheduleFrame, text = b1, borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                                                                column = empDispCol + 2,
                                                                                                                sticky = 'nsew')
                    Label(scheduleFrame, text = b2, borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                                                               column = empDispCol + 3,
                                                                                                               sticky = 'nsew')
                    Label(scheduleFrame, text = b3, borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                                                                column = empDispCol + 4,
                                                                                                                sticky = 'nsew')
                elif empObj.name != 'placeholder':
                    breakObjs[dept].assignBreaks(empObj)
                    
                    empDept = empObj.getDepartment()
                    deptClass = departments.get(empDept)
                    empInd = empObj.index
                    empDispRow = deptClass.dispRow
                    empDispCol = deptClass.dispCol ##doing it this way so backups are displayed on their own row
                    
                    if empObj.specialAttribute == 'Clean':
                        empColor = '#BCF3F9'
                    else:
                        empColor = '#F0F0F0'
                        
                    if empObj.getRestBreak1() == 0:
                        RB1 = "";
                    elif empObj.getRestBreak1() == -1:
                        RB1 = "Couldn't Find Break"
                    else:
                        RB1 = empObj.getRestBreak1()
                        
                    if empObj.getMealBreak() == 0:
                        MB = "";
                    elif empObj.getMealBreak() == -1:
                        MB = "Couldn't Find Break"
                    else:
                        MB = empObj.getMealBreak()
                        
                    if empObj.getRestBreak2() == 0:
                        RB2 = "";
                    elif empObj.getRestBreak2() == -1:
                        RB2 = "Couldn't Find Break"
                    else:
                        RB2 = empObj.getRestBreak2()
                        
                    
                    Label(scheduleFrame, text = empObj.getName(), borderwidth = 1, relief = 'solid', bg = empColor).grid(row = empInd + empDispRow + 1,
                                                                                                          column = empDispCol,
                                                                                                          sticky = 'nsew')
                    Label(scheduleFrame, text = str(toHM(empObj.getStartTime())) + "-"+ str(toHM(empObj.getEndTime())),
                          borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                  column = empDispCol + 1,
                                                                  sticky = 'nsew')
                    Label(scheduleFrame, text = RB1, borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                                                                column = empDispCol + 2,
                                                                                                                sticky = 'nsew')
                    Label(scheduleFrame, text = MB, borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                                                               column = empDispCol + 3,
                                                                                                               sticky = 'nsew')
                    Label(scheduleFrame, text = RB2, borderwidth = 1, relief = 'solid').grid(row = empInd + empDispRow + 1,
                                                                                                                column = empDispCol + 4,
                                                                                                                sticky = 'nsew')

    for i in range(len(breakObjs)):
        breakObjs[i].emps.clear()
        breakObjs[i].resetBreaks()
        
    empObjs.clear()

    scheduleFrame.grid(row = 1, column = 1, sticky = 'nsew', padx = 10)
    
                   
#################################
##---------CREDIT SETUP--------##
#################################
    
Label(root, text = "Built by Mikayla Billings-Alston, https://github.com/mcbillings/", font = ("Arial", 11)).grid(row = 3, column = 1, sticky = 'se')               

